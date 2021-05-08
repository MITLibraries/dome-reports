import getopt
import datetime
import calendar
import logging
import logging.config
import sqlite3
import pandas as pd
import numpy as np

from sys import argv, exit
#from xhtml2pdf import pisa
from contextlib import closing

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, Border, Side

'''
Originally this script only queried the SQLite db for item counts per
collection, and then produced reports for this in Pandas as specified.

Months later, the special case of a collection with no reports came up.
Since the query results used to populate the SQLite db did not pull
rows for empty collections, the reports here did not include said empty
collections.

As of now (5/7/21) the chosen solution is to add in rows for the empty
collections based on the complete list of collections in the  SQLite db
vs. the list of collections in the item count.  In the pandas pivot 
table, only a single instance of the empty collection is necessary for
the reports to display properly.

This notion of an 'empty' collection is not entirely accurate, as 
withdrawn and not-in-archive items are excluded in the Postgres query.
Linked items are not included either.  But this solution does
clarify the existence of a collection for which items are not being counted.

'''
def main(argv):

    db_filepath = "drp.db"

    # month is used only in the names of the report files
    # or, in the case of empty collections, it is the default month
    month = year = 0
    fmts = 'a'
    output_dir = "./reports"

    ASCII = True  #default
    CSV = False
    HTML = False
    MARKDOWN = False
    XLSX = False
    SCREEN = False 

    logging.config.fileConfig('logs/drplog2.config')
    logging.info("start of report creation")

    # Commandline options
    try:
        opts, args = getopt.getopt(argv,"hy:f:d:o:",
                  ["year=","fmts=","db=","output_dir="])
    except getopt.GetoptError:
        print_help()
        exit(2)

    for opt, arg in opts:
        if opt in ("-h", "--help"):
            print_help()
            exit()
        elif opt in ("-y", "--year"):
            year = int(arg)
        elif opt in ("-f", "--fmts"):
            fmts = arg
            ASCII = ('a' in fmts)
            CSV = ('c' in fmts)
            HTML = ('h' in fmts)
            MARKDOWN = ('m' in fmts)
            SCREEN = ('s' in fmts)
            XLSX = ('x' in fmts)
        elif opt in ("-d", "--testdb"):
            db_filepath = arg
        elif opt in ("-o", "--output_dir"):
            output_dir = arg

    today = datetime.date.today()

    if year == 0:
       year = today.year
       month = today.month

    logging.info(f"Creating Collection Item Count report for {year}")
    logging.info(f"Using database file {db_filepath}")

    queryCounts = ("SELECT comm.short_name as Community, "
          "coll.short_name as Collection, month, item_count "
          "FROM Community comm "
          "JOIN Collection coll ON comm.uuid = coll.comm_uuid "
          "JOIN Monthly_Item_Count itc ON coll.uuid = itc.coll_uuid "
          "WHERE year = ? AND coll.reportable = 1 "
          "ORDER BY coll.comm_uuid, coll.name;")

    queryColls = ("SELECT comm.short_name  as Community, "
          "coll.short_name as Collection "
          "FROM Community comm "
          "JOIN Collection coll ON comm.uuid = coll.comm_uuid "
          "WHERE coll.reportable = 1;")

    rows = None
    coll_rows = None

    try:
        with closing(sqlite3.connect(db_filepath)) as conn:
            with closing(conn.cursor()) as cursor:
                rows = cursor.execute(queryCounts, (year,)).fetchall()
                coll_rows = cursor.execute(queryColls,).fetchall()
    except sqlite3.Warning as w:
        logging.warning("SQLite Warning: " + str(w))
    except sqlite3.Error as e:
        logging.error("SQLite Error: " + str(e))
        exit(2)

    if len(rows) == 0:
        logging.error(f"No data in {db_filepath} for year {year}")
        exit(1)
    else:
        logging.debug(f'count of item count data points: {len(rows)}')

    if len(coll_rows) == 0:
        logging.error(f"Error retrieving collections from db")
        exit(1)
    else:
        logging.debug(f'count of all collections: {len(coll_rows)}')

    dfColl = pd.DataFrame(coll_rows)
    dfColl.columns = ["Community", "Collection"]
    logging.info(f"Count of all collections in the db: {len(dfColl)}")

    df = pd.DataFrame(rows)
    df.columns = ["Community", "Collection", "Month", "Count"]

    dfCollsWithCounts = df[["Community", "Collection"]].drop_duplicates()
    dfCollsWithCounts.columns = ["Community", "Collection"]

    if not len(dfCollsWithCounts) == len(dfColl):
        # https://cmsdk.com/python/pandas-get-rows-which-are-not-in-other-dataframe.html
        # remove non-empty collections from dfColl, i.e. the colls "with counts"
        dfColl = dfColl[~dfColl.isin(dfCollsWithCounts.to_dict('list')).all(1)]
        dfColl['Month'] = month
        dfColl['Count'] = 0
        df = df.append(dfColl)
        logging.info(f"added rows for {len(dfColl)} empty collections")

    #pt = df.pivot_table( index=[0], values=[2], columns=[1], aggfunc=max)
    pt = df.pivot_table( index=["Community", "Collection"],
              values=["Count"], columns=["Month"], fill_value=0, aggfunc=max)

    #extract column integers (will these always be sorted?)
    #i.e. the exact months in query result, skipping months with no data
    ar_mo_int = [ j for i, j in pt.columns.to_flat_index().to_list()]
    #get array of corresponding months
    pt.columns = [calendar.month_abbr[i] for i in ar_mo_int]

    #add column totals
    sums =  pt.sum(numeric_only=True, axis=0)   #this is a Series
    sums.name = ("Totals", "")
    pt = pt.append(sums)

    rpt_filepath = f'{output_dir}/drp_col_it_{year}-{month}'

    if ASCII:
        with open(f'{rpt_filepath}.txt', 'w') as ascii_out:
            table = pt.to_string(index = True)
            ascii_out.write(add_common_header(table, year))
            logging.info(f"created: {rpt_filepath}.txt") 
    if CSV:  #no headers??
        pt.to_csv(f'{rpt_filepath}.csv', index = True)
        logging.info(f"created: {rpt_filepath}.csv") 
    if HTML:
        with open(f'{rpt_filepath}.html', 'w') as html_out:
            html = pt.to_html().replace(" style=\"text-align: right;\"", "", 1)
            html_out.write(format_html(html, year))
        logging.info(f"created: {rpt_filepath}.html") 
    if MARKDOWN:
        with open(f'{rpt_filepath}.md', 'w') as md_out:
            table = pt.to_markdown(index = True)
            md_out.write(format_markdown(table, year))
            logging.info(f"created: {rpt_filepath}.md") 
    if XLSX:
        # style_xlsx(pt, f'{rpt_filepath}.xlsx')

        pt.to_excel(f'{rpt_filepath}.xlsx')   #, index = True)
        logging.info(f"created: {rpt_filepath}.xlsx") 

        style_xlsx(f'{rpt_filepath}.xlsx', len(ar_mo_int), len(pt) + 1)
        logging.info(f'restyled excel report {rpt_filepath}.xlsx')

    # print to console
    if SCREEN:
        print(add_common_header(pt.to_string(index = True), year))

    logging.info("Report creation processing completed\n")
    logging.shutdown()

def print_help():
    print('Usage: python3 create_rpts.py -h(elp) -y <year> -f <formats>' \
          ' -d <database> -o <output dir> -s <console output>')
    print('with formats: a = ascii (default), c = csv, h = html, ' \
          'm = markdown, x = xlsx (excel), s = screen')
    print('The default year is the current year ')
    print('')

"""  currently not used
pdf = open("drp2.pdf", "w+b")
pisa_status = pisa.CreatePDF(html, dest=pdf)
pdf.close()
print("pisa_status: " + str(pisa_status.err));
"""

def add_common_header(table, year):
    return f"""\n    MIT Libraries  -- dome.mit.edu\n
    Monthly Reports -- {year}\n    Collection Item Counts\n\n""" \
    + table + "\n"

def style_xlsx(excel_filename, ncols, nrows):


    c_style = NamedStyle(name="c_style")
    c_style.font = Font(bold=False)
    c_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = load_workbook(filename = excel_filename)
    ws = wb.active

    ws.column_dimensions['A'].width = 31
    ws.column_dimensions['B'].width = 38

    for col in ["A", "B"]:
        for cell in ws[col]:
            cell.style = c_style

    # put the first two col headers back to bold
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    # add border line above totals
    month_cols = [chr(x) for x in range(67, 67+ncols )]
    cells = [month_cols[i] + str(nrows) for i in range(0,ncols)]
    double = Side(border_style="medium", color="000000")

    for cell in cells:
        ws[cell].border = Border(top=double)

    wb.save(filename=excel_filename)

# community and collection are presented tuple
# tabulate and panda docs don't offer an option to break up as columns
# may be better to write a custom formatter here
def format_markdown(table, year):
    return "# MIT Libraries  -- dome.mit.edu" \
        + f"\n## Monthly Reports -- {year}\n## Collection Item Counts\n\n" \
        + table + "\n"


def format_html(table, year):

    template_top = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <title>dome.mit.edu - Collection Item Counts Monthly Report</title>

    <style>
    body {{
    font-family: Garamond, Arial;
    }}

    h1 {{
    font-family: Garamond, Arial;
    font-weight: normal;
    margin-left: 40px;
    }}

    table {{
    border-collapse: collapse;
    margin-left: 50px;
    }}

    table tbody tr th {{
    font-family: "Times New Roman", Times;
    text-align: left;
    padding: 0px 5px;
    }}

    table tbody tr td {{
    text-align: right;
    padding: 0px 3px;
    }}

    </style>
    </head>
    <body>

    <h2>MIT Libraries  -- dome.mit.edu</h2>
    <h1>Monthly Reports -- {year}</h1>
    <h1>Collection Item Counts</h1>

    """

    template_bottom = """
    </body>
    </html>

    """

    return template_top + table + template_bottom

main((argv[1:]))
