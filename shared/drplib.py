import logging
import pandas as pd

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, Border, Side


def cleanup_and_exit(ret : int):
    logging.shutdown()
    exit(ret)

 # Check to avoid redundant processing
def datafile_processed(cursor, filepath):
    c = cursor.execute("SELECT COUNT(*) FROM FilesProcessed WHERE name = ?",
        (filepath.name, )).fetchall()
    if c[0][0] > 0:
        #logging.warning(f"{filepath.name} already processed.")
        return True
    logging.debug(f"Verified file {filepath.name} not previously processed.")
    return False


# For string output of a report, add a header with a date
def add_common_header(table, year, datestring):
    return f"""\n    MIT Libraries  -- dome.mit.edu\n
    Dome Reports -- {datestring}\n
    Monthly Collection Item Counts Time Series for {year}\n\n""" \
        + table + "\n"

#param 'year' is for the period reported on
def write_rpts(df, fmts, dirpath, stem, year):


    datestring = datetime.now().isoformat()[:10]

    if fmts.ASCII:
        rpt_filepath = dirpath / (stem + '.txt')
        with open(rpt_filepath, 'w') as ascii_out:
            table = df.to_string(index = True)
            ascii_out.write(add_common_header(table, year, datestring))
            logging.info(f"created: {rpt_filepath.name}") 

    #CSV format includes headers
    #the field separator is a comma
    if fmts.CSV:
        rpt_filepath = dirpath / (stem + '.csv')
        df.to_csv(rpt_filepath, index = True)
        logging.info(f"created: {rpt_filepath.name}") 

    if fmts.HTML:
        rpt_filepath = dirpath / (stem + '.html')
        with open(rpt_filepath, 'w') as html_out:
            #remove unwanted text alignment
            html = df.to_html().replace(" style=\"text-align: right;\"", "", 1)
            html_out.write(format_html(html, year))
        logging.info(f"created: {rpt_filepath.name}")

    if fmts.MARKDOWN:
        rpt_filepath = dirpath / (stem + '.md')
        with open(rpt_filepath, 'w') as md_out:
            table = df.to_markdown(index = True)
            md_out.write(format_markdown(table))
            logging.info(f"created: {rpt_filepath.name}")

    if fmts.XLSX:
        rpt_filepath = dirpath / (stem + '.xlsx')

        # must style after writing out the file
        df.to_excel(rpt_filepath)   #, index = True)
        logging.info(f"created report: {rpt_filepath.name}") 

        style_xlsx(rpt_filepath, df.shape[1], df.shape[0])
        logging.debug(f'restyled excel report {rpt_filepath.name}')

    # print to console
    if fmts.SCREEN:
        print(add_common_header(df.to_string(index = True), year, datestring))

# This styling may not generalize to future reports
def style_xlsx(excel_filename, ncols, nrows):

    c_style = NamedStyle(name="c_style")
    c_style.font = Font(bold=False)
    c_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = load_workbook(filename = excel_filename)
    ws = wb.active

    ws.column_dimensions['A'].width = 31
    ws.column_dimensions['B'].width = 38

    if ncols == 1:
        ws.column_dimensions['C'].width = 15

    for col in ["A", "B"]:
        for cell in ws[col]:
            cell.style = c_style

    # put the first two col headers back to bold
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    # add border line above totals
    month_cols = [chr(x) for x in range(67, 67+ncols )]
    cells = [month_cols[i] + str(nrows+1) for i in range(0,ncols)]
    double = Side(border_style="medium", color="000000")

    for cell in cells:
        ws[cell].border = Border(top=double)

    wb.save(filename=excel_filename)


# community and collection are presented tuple
# tabulate and panda docs don't offer an option to break up as columns
# may be better to write a custom formatter here
def format_markdown(table, datestring):
    return "# MIT Libraries  -- dome.mit.edu" \
        + f"\n## Monthly Reports -- {datestring}\n## Collection Item Counts\n\n" \
        + table + "\n"


def format_html(table, year):

    template_top = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <title>dome.mit.edu - Collection Item Counts Monthly Report</title>

    <style>
    body {{
    font-family: Garamond, Arial;
    }}

    h1 {{
    font-family: Garamond, Arial;
    font-weight: normal;
    margin-left: 40px;
    }}

    table {{
    border-collapse: collapse;
    margin-left: 50px;
    }}

    table tbody tr th {{
    font-family: "Times New Roman", Times;
    text-align: left;
    padding: 0px 5px;
    }}

    table tbody tr td {{
    text-align: right;
    padding: 0px 3px;
    }}

    </style>
    </head>
    <body>

    <h2>MIT Libraries  -- dome.mit.edu</h2>
    <h1>Collection Item Counts for {year}</h1>

    """

    template_bottom = """
    </body>
    </html>

    """

    return template_top + table + template_bottom

